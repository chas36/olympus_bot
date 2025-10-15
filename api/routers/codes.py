from fastapi import APIRouter, UploadFile, File, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_, Integer
from typing import List, Dict
import tempfile
import os
import io
import csv
import logging

from database.database import get_async_session
from database.models import OlympiadSession, Grade8Code, Grade9Code, Student, OlympiadCode, Grade8ReserveCode, moscow_now
from parser.csv_parser import parse_codes_csv
from datetime import datetime

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/codes", tags=["Codes"])


@router.post("/upload-csv")
async def upload_codes_csv(
    files: List[UploadFile] = File(...),
    auto_reserve: bool = Query(True, description="Автоматически резервировать коды 9 класса для 8"),
    session: AsyncSession = Depends(get_async_session)
):
    """
    Загрузка кодов из CSV с группировкой по предметам.

    Все коды одного предмета из разных параллелей объединяются в одну сессию.
    Дата берется из CSV файла.
    """
    results = []
    # Словарь для группировки: {subject: {date: datetime, codes_by_class: {8: [...], 9: [...]}}}
    subjects_map = {}

    for file in files:
        if not file.filename.endswith('.csv'):
            continue

        # Сохраняем временно
        with tempfile.NamedTemporaryFile(delete=False, suffix='.csv') as tmp:
            content = await file.read()
            tmp.write(content)
            tmp_path = tmp.name

        try:
            # Парсим - пробуем сначала utf-8, потом windows-1251
            try:
                parsed = parse_codes_csv(tmp_path, encoding='utf-8')
            except UnicodeDecodeError:
                parsed = parse_codes_csv(tmp_path, encoding='windows-1251')

            for subject_data in parsed:
                subject = subject_data['subject']
                class_num = subject_data['class_number']
                codes = subject_data['codes']
                date = subject_data.get('date') or moscow_now()

                # Группируем по предметам
                if subject not in subjects_map:
                    subjects_map[subject] = {
                        'date': date,
                        'codes_by_class': {}
                    }

                # Добавляем коды для этого класса
                if class_num not in subjects_map[subject]['codes_by_class']:
                    subjects_map[subject]['codes_by_class'][class_num] = []

                subjects_map[subject]['codes_by_class'][class_num].extend(codes)

                results.append({
                    "file": file.filename,
                    "subject": subject,
                    "class": class_num,
                    "codes_count": len(codes)
                })

        finally:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)

    # Создаем сессии для каждого предмета
    created_sessions = []
    for subject, data in subjects_map.items():
        # Проверяем, существует ли уже сессия для этого предмета
        result = await session.execute(
            select(OlympiadSession).where(
                OlympiadSession.subject == subject
            )
        )
        existing_session = result.scalar_one_or_none()

        if existing_session:
            # Обновляем существующую сессию
            olympiad = existing_session
            olympiad.date = data['date']
        else:
            # Создаем новую сессию
            olympiad = OlympiadSession(
                subject=subject,
                date=data['date'],
                is_active=False
            )
            session.add(olympiad)

        await session.flush()

        # Добавляем коды всех классов (универсальная система 5-11)
        for class_num, codes in data['codes_by_class'].items():
            for code_str in codes:
                code = OlympiadCode(
                    session_id=olympiad.id,
                    class_number=class_num,
                    code=code_str,
                    is_assigned=False,
                    is_issued=False
                )
                session.add(code)

        created_sessions.append({
            "subject": subject,
            "date": data['date'].isoformat() if isinstance(data['date'], datetime) else str(data['date']),
            "classes": list(data['codes_by_class'].keys())
        })

    await session.commit()

    # Автоматическое резервирование
    if auto_reserve:
        reserve_result = await reserve_grade9_for_grade8(session)
        return {
            "success": True,
            "uploaded": results,
            "sessions_created": created_sessions,
            "reservation": reserve_result
        }

    return {
        "success": True,
        "uploaded": results,
        "sessions_created": created_sessions
    }


async def reserve_grade9_for_grade8(session: AsyncSession, session_id: int = None) -> Dict:
    """
    Автоматическое резервирование кодов 9 класса для 8 классов

    Логика:
    1. Группируем 8-классников по параллелям
    2. Для каждой параллели выделяем количество резервных кодов = количеству учеников
    3. Берем нераспределенные коды 9 класса

    Args:
        session: Сессия БД
        session_id: ID сессии олимпиады (если None, используется активная сессия)
    """
    # Получаем сессию олимпиады
    if session_id:
        result = await session.execute(
            select(OlympiadSession).where(OlympiadSession.id == session_id)
        )
        active_session = result.scalar_one_or_none()
        if not active_session:
            return {"message": f"Сессия {session_id} не найдена", "reserved": 0}
    else:
        result = await session.execute(
            select(OlympiadSession).where(OlympiadSession.is_active == True)
        )
        active_session = result.scalar_one_or_none()
        if not active_session:
            return {"message": "Нет активной сессии", "reserved": 0}

    # Получаем 8-классников по параллелям
    result = await session.execute(
        select(Student).where(Student.class_number == 8)
    )
    grade8_students = result.scalars().all()

    if not grade8_students:
        return {"message": "Нет учеников 8 класса", "reserved": 0}

    # Группируем по параллелям
    from collections import defaultdict
    students_by_parallel = defaultdict(list)
    for student in grade8_students:
        parallel = f"8{student.parallel or ''}"
        students_by_parallel[parallel].append(student)

    # Получаем нераспределенные коды 9 класса
    result = await session.execute(
        select(OlympiadCode).where(
            and_(
                OlympiadCode.session_id == active_session.id,
                OlympiadCode.class_number == 9,
                OlympiadCode.is_assigned == False
            )
        )
    )
    grade9_codes = result.scalars().all()

    if not grade9_codes:
        return {"message": "Нет доступных кодов 9 класса", "reserved": 0}

    # Считаем сколько учеников 9 класса
    result = await session.execute(
        select(func.count(Student.id)).where(Student.class_number == 9)
    )
    grade9_students_count = result.scalar()

    # Реальный резерв = нераспределенные коды минус количество учеников 9 класса
    # (т.к. эти коды должны быть зарезервированы для самих 9-классников)
    actual_surplus = len(grade9_codes) - grade9_students_count

    if actual_surplus <= 0:
        return {
            "message": f"Нет избыточных кодов 9 класса (всего кодов: {len(grade9_codes)}, учеников: {grade9_students_count})",
            "reserved": 0
        }

    logger.info(f"Всего кодов 9 класса: {len(grade9_codes)}, учеников 9 класса: {grade9_students_count}, избыток: {actual_surplus}")

    # Берем только избыточные коды для резервирования
    grade9_codes_for_reserve = grade9_codes[:actual_surplus]

    # Распределяем коды пропорционально численности параллелей
    total_students = sum(len(students) for students in students_by_parallel.values())
    total_available_codes = len(grade9_codes_for_reserve)

    logger.info(f"Всего учеников 8 класса: {total_students}")
    logger.info(f"Доступных кодов 9 класса: {total_available_codes}")

    # Рассчитываем пропорции для каждой параллели
    distribution = {}
    for parallel, students in students_by_parallel.items():
        student_count = len(students)
        proportion = student_count / total_students
        codes_for_parallel = int(total_available_codes * proportion)
        distribution[parallel] = {
            'student_count': student_count,
            'proportion': proportion,
            'codes_allocated': codes_for_parallel
        }

    # Распределяем коды
    code_index = 0
    total_reserved = 0

    for parallel in sorted(distribution.keys()):
        info = distribution[parallel]
        codes_to_allocate = info['codes_allocated']

        for i in range(codes_to_allocate):
            if code_index >= len(grade9_codes_for_reserve):
                break

            reserve_code = Grade8ReserveCode(
                session_id=active_session.id,
                class_parallel=parallel,
                code=grade9_codes_for_reserve[code_index].code
            )
            session.add(reserve_code)
            code_index += 1
            total_reserved += 1

        logger.info(f"Для {parallel}: {info['student_count']} учеников, выделено {codes_to_allocate} резервных кодов ({info['proportion']*100:.1f}%)")

    await session.commit()

    return {
        "message": f"Зарезервировано {total_reserved} кодов из 9 класса для 8 классов (пропорционально численности)",
        "reserved": total_reserved,
        "parallels": len(students_by_parallel),
        "distribution": {k: v['codes_allocated'] for k, v in distribution.items()}
    }


@router.post("/reserve")
async def manual_reserve(
    session: AsyncSession = Depends(get_async_session)
):
    """Ручное резервирование кодов 9→8"""
    result = await reserve_grade9_for_grade8(session)
    return result


@router.get("/sessions")
async def get_sessions(
    session: AsyncSession = Depends(get_async_session)
):
    """Получить все сессии"""
    result = await session.execute(
        select(OlympiadSession).order_by(OlympiadSession.date.desc())
    )
    sessions = result.scalars().all()
    
    return [
        {
            "id": s.id,
            "subject": s.subject,
            "date": s.date.isoformat(),
            "is_active": s.is_active
        }
        for s in sessions
    ]


@router.post("/sessions/{session_id}/activate")
async def activate_session(
    session_id: int,
    session: AsyncSession = Depends(get_async_session)
):
    """Активировать сессию"""
    # Деактивируем все
    result = await session.execute(select(OlympiadSession))
    for s in result.scalars().all():
        s.is_active = False
    
    # Активируем нужную
    result = await session.execute(
        select(OlympiadSession).where(OlympiadSession.id == session_id)
    )
    target = result.scalar_one_or_none()
    
    if not target:
        raise HTTPException(404, "Сессия не найдена")
    
    target.is_active = True
    await session.commit()
    
    return {"success": True, "session_id": session_id}


@router.get("/stats")
async def get_codes_stats(
    session: AsyncSession = Depends(get_async_session)
):
    """Статистика по кодам для всех классов (5-11)"""
    # Получаем статистику по каждому классу
    result = await session.execute(
        select(
            OlympiadCode.class_number,
            func.count(OlympiadCode.id).label('total'),
            func.sum(func.cast(OlympiadCode.is_assigned, Integer)).label('assigned'),
            func.sum(func.cast(OlympiadCode.is_issued, Integer)).label('issued')
        )
        .group_by(OlympiadCode.class_number)
        .order_by(OlympiadCode.class_number)
    )

    stats_by_class = {}
    for row in result.fetchall():
        class_num, total, assigned, issued = row
        stats_by_class[f"grade{class_num}"] = {
            "class": class_num,
            "total": total or 0,
            "assigned": assigned or 0,
            "unassigned": (total or 0) - (assigned or 0),
            "issued": issued or 0
        }

    # Статистика по резервным кодам для 8 класса
    result = await session.execute(
        select(
            func.count(Grade8ReserveCode.id).label('total'),
            func.sum(func.cast(Grade8ReserveCode.is_used, Integer)).label('used')
        )
    )
    reserve_row = result.first()

    reserve_stats = {
        "total": reserve_row.total or 0,
        "used": reserve_row.used or 0,
        "available": (reserve_row.total or 0) - (reserve_row.used or 0)
    }

    return {
        "by_class": stats_by_class,
        "grade8_reserve": reserve_stats
    }


@router.get("/export/session/{session_id}")
async def export_session_codes(
    session_id: int,
    session: AsyncSession = Depends(get_async_session)
):
    """Экспорт кодов сессии в виде ZIP-архива с Excel файлами по классам и параллелям"""
    from urllib.parse import quote
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    import zipfile

    # Получаем сессию
    result = await session.execute(
        select(OlympiadSession).where(OlympiadSession.id == session_id)
    )
    olympiad = result.scalar_one_or_none()

    if not olympiad:
        raise HTTPException(404, "Сессия не найдена")

    # Получаем коды из универсальной таблицы с загрузкой связанных студентов
    from sqlalchemy.orm import joinedload

    result = await session.execute(
        select(OlympiadCode)
        .options(joinedload(OlympiadCode.student))
        .where(OlympiadCode.session_id == session_id)
        .order_by(OlympiadCode.class_number, OlympiadCode.student_id)
    )
    universal_codes = result.scalars().all()

    # Получаем резервные коды для 8 класса (из пула 9 класса)
    result = await session.execute(
        select(Grade8ReserveCode)
        .options(joinedload(Grade8ReserveCode.used_by))
        .where(Grade8ReserveCode.session_id == session_id)
        .order_by(Grade8ReserveCode.class_parallel)
    )
    reserve_codes = result.scalars().all()

    # Группируем коды по классам и параллелям
    codes_by_class_parallel = {}
    for code in universal_codes:
        class_key = code.class_number
        if class_key not in codes_by_class_parallel:
            codes_by_class_parallel[class_key] = {}

        # Определяем параллель из студента или пропускаем нераспределенные
        if code.student and code.student.parallel:
            parallel = code.student.parallel

            if parallel not in codes_by_class_parallel[class_key]:
                codes_by_class_parallel[class_key][parallel] = []

            codes_by_class_parallel[class_key][parallel].append(code)

    # Добавляем резервные коды для 8 класса
    reserve_by_parallel = {}
    for reserve_code in reserve_codes:
        # "8А" -> "А", "8И" -> "И"
        class_parallel = reserve_code.class_parallel
        if class_parallel.startswith('8'):
            parallel = class_parallel[1:]  # Убираем первый символ "8"
        else:
            parallel = class_parallel

        if parallel not in reserve_by_parallel:
            reserve_by_parallel[parallel] = []
        reserve_by_parallel[parallel].append(reserve_code)

    # Создаём временную директорию для файлов
    zip_buffer = io.BytesIO()

    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        # Для каждого класса создаём папку и файлы параллелей
        for class_num in sorted(codes_by_class_parallel.keys()):
            parallels = codes_by_class_parallel[class_num]

            for parallel in sorted(parallels.keys()):
                codes = parallels[parallel]

                # Создаём Excel файл для этой параллели
                wb = Workbook()
                ws = wb.active
                ws.title = f"{class_num}{parallel}"

                # Заголовок
                ws['A1'] = f"Коды для {class_num}{parallel} - {olympiad.subject}"
                ws['A1'].font = Font(bold=True, size=14)
                ws.merge_cells('A1:B1')
                ws['A1'].alignment = Alignment(horizontal='center')

                # Заголовки колонок
                ws['A3'] = "ФИО"
                ws['B3'] = "Код"
                ws['A3'].font = Font(bold=True)
                ws['B3'].font = Font(bold=True)
                ws['A3'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                ws['B3'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

                # Данные
                row = 4
                for code in codes:
                    if code.student:
                        ws[f'A{row}'] = code.student.full_name
                        ws[f'B{row}'] = code.code
                        row += 1

                # Добавляем резервные коды для 8 класса
                if class_num == 8 and parallel in reserve_by_parallel:
                    # Добавляем разделитель
                    if row > 4:  # Есть основные коды
                        row += 1
                        ws[f'A{row}'] = "РЕЗЕРВНЫЕ КОДЫ (из пула 9 класса)"
                        ws[f'A{row}'].font = Font(bold=True, italic=True)
                        ws.merge_cells(f'A{row}:B{row}')
                        ws[f'A{row}'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                        row += 1

                    # Добавляем резервные коды
                    for reserve_code in reserve_by_parallel[parallel]:
                        # Если код уже использован, показываем кто его получил
                        if reserve_code.is_used and reserve_code.used_by:
                            ws[f'A{row}'] = f"{reserve_code.used_by.full_name} (резервный)"
                            ws[f'A{row}'].font = Font(italic=True, color="0070C0")
                        else:
                            ws[f'A{row}'] = "Резервный код"
                        ws[f'B{row}'] = reserve_code.code
                        row += 1

                # Ширина колонок
                ws.column_dimensions['A'].width = 40
                ws.column_dimensions['B'].width = 30

                # Сохраняем во временный буфер
                excel_buffer = io.BytesIO()
                wb.save(excel_buffer)
                excel_buffer.seek(0)

                # Добавляем в архив
                folder_name = f"{class_num}_класс"
                file_name = f"{class_num}{parallel}.xlsx"
                zip_file.writestr(f"{folder_name}/{file_name}", excel_buffer.getvalue())

    zip_buffer.seek(0)

    # Используем транслитерацию для имени файла
    transliteration = {
        'а': 'a', 'б': 'b', 'в': 'v', 'г': 'g', 'д': 'd', 'е': 'e', 'ё': 'yo',
        'ж': 'zh', 'з': 'z', 'и': 'i', 'й': 'y', 'к': 'k', 'л': 'l', 'м': 'm',
        'н': 'n', 'о': 'o', 'п': 'p', 'р': 'r', 'с': 's', 'т': 't', 'у': 'u',
        'ф': 'f', 'х': 'h', 'ц': 'ts', 'ч': 'ch', 'ш': 'sh', 'щ': 'sch',
        'ъ': '', 'ы': 'y', 'ь': '', 'э': 'e', 'ю': 'yu', 'я': 'ya',
        ' ': '_', '-': '_'
    }

    safe_subject = olympiad.subject.lower()
    for ru, en in transliteration.items():
        safe_subject = safe_subject.replace(ru, en)

    date_str = olympiad.date.strftime('%d-%m-%Y')
    filename = f"{safe_subject}_{date_str}.zip"

    return StreamingResponse(
        iter([zip_buffer.getvalue()]),
        media_type="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"'
        }
    )


@router.post("/sessions/{session_id}/distribute")
async def distribute_codes_to_students(
    session_id: int,
    session: AsyncSession = Depends(get_async_session)
):
    """
    Распределение кодов сессии между всеми учениками по классам и параллелям

    Логика:
    1. Берём все нераспределённые коды для каждого класса
    2. Берём всех учеников этого класса (всех, не только зарегистрированных)
    3. Распределяем коды по ученикам каждой параллели
    """
    from sqlalchemy.orm import joinedload

    # Проверяем существование сессии
    result = await session.execute(
        select(OlympiadSession).where(OlympiadSession.id == session_id)
    )
    olympiad = result.scalar_one_or_none()

    if not olympiad:
        raise HTTPException(404, "Сессия не найдена")

    # Получаем все нераспределённые коды этой сессии, сгруппированные по классам
    result = await session.execute(
        select(OlympiadCode)
        .where(
            and_(
                OlympiadCode.session_id == session_id,
                OlympiadCode.is_assigned == False,
                OlympiadCode.student_id.is_(None)
            )
        )
        .order_by(OlympiadCode.class_number, OlympiadCode.id)
    )
    available_codes = result.scalars().all()

    if not available_codes:
        return {
            "success": True,
            "message": "Нет доступных кодов для распределения",
            "distributed": 0
        }

    # Группируем коды по классам
    codes_by_class = {}
    for code in available_codes:
        if code.class_number not in codes_by_class:
            codes_by_class[code.class_number] = []
        codes_by_class[code.class_number].append(code)

    distributed_count = 0
    distribution_log = []

    # Для каждого класса распределяем коды
    for class_num, codes in codes_by_class.items():
        # Получаем всех учеников этого класса, сгруппированных по параллелям
        result = await session.execute(
            select(Student)
            .where(Student.class_number == class_num)
            .order_by(Student.parallel, Student.full_name)
        )
        students = result.scalars().all()

        if not students:
            logger.warning(f"Нет учеников для класса {class_num}")
            continue

        # Группируем учеников по параллелям
        students_by_parallel = {}
        for student in students:
            parallel = student.parallel or "Без параллели"
            if parallel not in students_by_parallel:
                students_by_parallel[parallel] = []
            students_by_parallel[parallel].append(student)

        # Распределяем коды по параллелям
        code_index = 0
        for parallel, parallel_students in sorted(students_by_parallel.items()):
            parallel_distributed = 0

            for student in parallel_students:
                if code_index >= len(codes):
                    logger.warning(f"Закончились коды для класса {class_num}")
                    break

                code = codes[code_index]
                code.student_id = student.id
                code.is_assigned = True
                code.assigned_at = moscow_now()

                code_index += 1
                distributed_count += 1
                parallel_distributed += 1

            distribution_log.append({
                "class": f"{class_num}{parallel}",
                "students": len(parallel_students),
                "codes_assigned": parallel_distributed
            })

            if code_index >= len(codes):
                break

    await session.commit()

    # Автоматически резервируем коды 9 класса для 8 класса
    reserve_result = await reserve_grade9_for_grade8(session, session_id=session_id)
    logger.info(f"Резервирование для сессии {session_id}: {reserve_result}")

    return {
        "success": True,
        "message": f"Распределено {distributed_count} кодов",
        "distributed": distributed_count,
        "distribution_log": distribution_log,
        "reserve_info": reserve_result
    }