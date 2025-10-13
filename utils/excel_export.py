"""
Модуль для экспорта данных в Excel
"""
import io
from typing import List
from datetime import datetime
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExcelExporter:
    """Класс для экспорта данных в Excel"""

    @staticmethod
    def export_students(students: List[dict]) -> io.BytesIO:
        """
        Экспортирует список учеников в Excel

        Args:
            students: Список словарей с данными учеников

        Returns:
            BytesIO объект с Excel файлом
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl не установлен. Установите: pip install openpyxl")

        wb = Workbook()
        ws = wb.active
        ws.title = "Ученики"

        # Заголовки
        headers = ["ID", "ФИО", "Класс", "Параллель", "Код регистрации", "Зарегистрирован", "Telegram ID", "Дата создания", "Дата регистрации"]

        # Стилизация заголовков
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Данные
        for row, student in enumerate(students, 2):
            ws.cell(row=row, column=1, value=student.get("id"))
            ws.cell(row=row, column=2, value=student.get("full_name"))
            ws.cell(row=row, column=3, value=student.get("class_number"))
            ws.cell(row=row, column=4, value=student.get("parallel"))
            ws.cell(row=row, column=5, value=student.get("registration_code"))
            ws.cell(row=row, column=6, value="Да" if student.get("is_registered") else "Нет")
            ws.cell(row=row, column=7, value=student.get("telegram_id") or "-")

            # Форматирование дат
            created_at = student.get("created_at")
            if created_at and isinstance(created_at, str):
                created_at = datetime.fromisoformat(created_at).strftime("%d.%m.%Y %H:%M")
            ws.cell(row=row, column=8, value=created_at)

            registered_at = student.get("registered_at")
            if registered_at and isinstance(registered_at, str):
                registered_at = datetime.fromisoformat(registered_at).strftime("%d.%m.%Y %H:%M")
            ws.cell(row=row, column=9, value=registered_at or "-")

        # Автоматическая ширина колонок
        for col in range(1, len(headers) + 1):
            max_length = 0
            column = get_column_letter(col)

            for cell in ws[column]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        # Сохранение в BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output

    @staticmethod
    def export_olympiads(olympiads: List[dict]) -> io.BytesIO:
        """Экспортирует список олимпиад в Excel"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl не установлен. Установите: pip install openpyxl")

        wb = Workbook()
        ws = wb.active
        ws.title = "Олимпиады"

        # Заголовки
        headers = ["ID", "Предмет", "Класс", "Дата проведения", "Этап", "Активна", "Файл", "Дата загрузки"]

        # Стилизация заголовков
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Данные
        for row, olympiad in enumerate(olympiads, 2):
            ws.cell(row=row, column=1, value=olympiad.get("id"))
            ws.cell(row=row, column=2, value=olympiad.get("subject"))
            ws.cell(row=row, column=3, value=olympiad.get("class_number") or "Разные")

            # Форматирование даты
            date = olympiad.get("date")
            if date and isinstance(date, str):
                date = datetime.fromisoformat(date).strftime("%d.%m.%Y")
            ws.cell(row=row, column=4, value=date)

            ws.cell(row=row, column=5, value=olympiad.get("stage") or "-")
            ws.cell(row=row, column=6, value="Да" if olympiad.get("is_active") else "Нет")
            ws.cell(row=row, column=7, value=olympiad.get("uploaded_file_name") or "-")

            # Дата загрузки
            upload_time = olympiad.get("upload_time")
            if upload_time and isinstance(upload_time, str):
                upload_time = datetime.fromisoformat(upload_time).strftime("%d.%m.%Y %H:%M")
            ws.cell(row=row, column=8, value=upload_time)

        # Автоматическая ширина колонок
        for col in range(1, len(headers) + 1):
            max_length = 0
            column = get_column_letter(col)

            for cell in ws[column]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        # Сохранение в BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output

    @staticmethod
    def export_statistics(stats: dict) -> io.BytesIO:
        """Экспортирует статистику в Excel"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl не установлен. Установите: pip install openpyxl")

        wb = Workbook()

        # Лист общей статистики
        ws1 = wb.active
        ws1.title = "Общая статистика"

        ws1["A1"] = "Параметр"
        ws1["B1"] = "Значение"

        ws1["A1"].font = Font(bold=True)
        ws1["B1"].font = Font(bold=True)

        row = 2
        for key, value in stats.get("general", {}).items():
            ws1.cell(row=row, column=1, value=key)
            ws1.cell(row=row, column=2, value=value)
            row += 1

        # Автоширина
        ws1.column_dimensions["A"].width = 40
        ws1.column_dimensions["B"].width = 20

        # Лист по классам
        if "classes" in stats:
            ws2 = wb.create_sheet("По классам")
            ws2["A1"] = "Класс"
            ws2["B1"] = "Всего учеников"
            ws2["C1"] = "Зарегистрировано"

            for col in ["A1", "B1", "C1"]:
                ws2[col].font = Font(bold=True)

            row = 2
            for class_data in stats["classes"]:
                ws2.cell(row=row, column=1, value=f"{class_data['class_number']} класс")
                ws2.cell(row=row, column=2, value=class_data.get("total", 0))
                ws2.cell(row=row, column=3, value=class_data.get("registered", 0))
                row += 1

            for col in ["A", "B", "C"]:
                ws2.column_dimensions[col].width = 20

        # Сохранение
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output
