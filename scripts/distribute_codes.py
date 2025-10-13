"""
Скрипт распределения кодов между учениками (Вариант 1)

Функционал:
1. Равномерно распределяет коды между учениками всех классов (5-11)
2. Для 8 класса дополнительно выделяет коды из пула 9 класса
3. Генерирует отдельные Excel файлы для каждого класса
4. Каждый файл содержит:
   - Основную таблицу: ФИО - Код
   - Для 8 класса: дополнительная таблица с резервными кодами

Использование:
    python scripts/distribute_codes.py --session-id 1
    python scripts/distribute_codes.py --session-id 1 --output-dir exports/
"""

import asyncio
import argparse
import os
from datetime import datetime
from collections import defaultdict
from sqlalchemy import select, func, and_
from sqlalchemy.ext.asyncio import AsyncSession

# Добавляем путь к проекту
import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from database.database import get_async_session
from database.models import Student, OlympiadSession, OlympiadCode, Grade8ReserveCode
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class CodeDistributor:
    """Класс для распределения кодов между учениками"""

    def __init__(self, session: AsyncSession, olympiad_session_id: int):
        self.session = session
        self.olympiad_session_id = olympiad_session_id
        self.stats = defaultdict(lambda: {
            'students': 0,
            'codes_assigned': 0,
            'codes_available': 0,
            'reserve_codes': 0
        })

    async def distribute_all(self):
        """Основной метод распределения кодов"""
        print(f"🚀 Начинаем распределение кодов для сессии {self.olympiad_session_id}")
        print("=" * 80)

        # Получаем информацию о сессии
        olympiad = await self.session.get(OlympiadSession, self.olympiad_session_id)
        if not olympiad:
            raise ValueError(f"Сессия {self.olympiad_session_id} не найдена")

        print(f"📚 Олимпиада: {olympiad.subject}")
        print(f"📅 Дата: {olympiad.date}")
        print()

        # Получаем всех учеников по классам
        students_by_class = await self._get_students_by_class()

        if not students_by_class:
            print("⚠️  Нет учеников в базе данных")
            return

        # Распределяем коды для каждого класса
        for class_number in sorted(students_by_class.keys()):
            await self._distribute_for_class(class_number, students_by_class[class_number])

        # Специальная обработка для 8 класса - резервные коды
        if 8 in students_by_class:
            await self._distribute_reserve_codes_for_grade8(students_by_class[8])

        # Коммитим все изменения
        await self.session.commit()

        # Выводим статистику
        self._print_statistics()

        print("\n✅ Распределение завершено!")

    async def _get_students_by_class(self):
        """Получить всех учеников, сгруппированных по классам"""
        result = await self.session.execute(
            select(Student)
            .where(Student.class_number.isnot(None))
            .order_by(Student.class_number, Student.parallel, Student.full_name)
        )
        students = result.scalars().all()

        students_by_class = defaultdict(list)
        for student in students:
            students_by_class[student.class_number].append(student)

        print(f"👥 Найдено учеников:")
        for class_num in sorted(students_by_class.keys()):
            print(f"   {class_num} класс: {len(students_by_class[class_num])} чел.")
        print()

        return students_by_class

    async def _distribute_for_class(self, class_number: int, students: list):
        """Распределить коды для одного класса"""
        print(f"📊 Обрабатываем {class_number} класс...")

        # Получаем доступные коды для этого класса
        result = await self.session.execute(
            select(OlympiadCode)
            .where(
                and_(
                    OlympiadCode.session_id == self.olympiad_session_id,
                    OlympiadCode.class_number == class_number,
                    OlympiadCode.is_assigned == False
                )
            )
            .order_by(OlympiadCode.id)
        )
        available_codes = result.scalars().all()

        if not available_codes:
            print(f"   ⚠️  Нет доступных кодов для {class_number} класса")
            self.stats[class_number]['students'] = len(students)
            return

        if len(available_codes) < len(students):
            print(f"   ⚠️  Недостаточно кодов! Нужно: {len(students)}, Доступно: {len(available_codes)}")

        # Распределяем коды между учениками
        assigned_count = 0
        for i, student in enumerate(students):
            if i >= len(available_codes):
                break

            code = available_codes[i]
            code.student_id = student.id
            code.is_assigned = True
            code.assigned_at = datetime.utcnow()
            assigned_count += 1

        # Обновляем статистику
        self.stats[class_number]['students'] = len(students)
        self.stats[class_number]['codes_assigned'] = assigned_count
        self.stats[class_number]['codes_available'] = len(available_codes) - assigned_count

        print(f"   ✅ Распределено: {assigned_count}/{len(students)}")

    async def _distribute_reserve_codes_for_grade8(self, grade8_students: list):
        """Распределить резервные коды из пула 9 класса для 8-классников"""
        print(f"\n🔄 Выделяем резервные коды 9 класса для 8 класса...")

        # Группируем 8-классников по параллелям
        students_by_parallel = defaultdict(list)
        for student in grade8_students:
            parallel = f"8{student.parallel or ''}"
            students_by_parallel[parallel].append(student)

        # Получаем нераспределенные коды 9 класса
        result = await self.session.execute(
            select(OlympiadCode)
            .where(
                and_(
                    OlympiadCode.session_id == self.olympiad_session_id,
                    OlympiadCode.class_number == 9,
                    OlympiadCode.is_assigned == False
                )
            )
            .order_by(OlympiadCode.id)
        )
        grade9_codes = result.scalars().all()

        if not grade9_codes:
            print("   ⚠️  Нет доступных кодов 9 класса для резерва")
            return

        print(f"   📦 Доступно кодов 9 класса: {len(grade9_codes)}")

        # Распределяем коды равномерно между параллелями 8 класса
        codes_per_parallel = len(grade9_codes) // len(students_by_parallel)
        total_allocated = 0

        code_index = 0
        for parallel, students in sorted(students_by_parallel.items()):
            # Количество кодов = количество учеников в параллели
            codes_needed = len(students)
            codes_to_allocate = min(codes_needed, codes_per_parallel, len(grade9_codes) - code_index)

            for i in range(codes_to_allocate):
                if code_index >= len(grade9_codes):
                    break

                reserve_code = Grade8ReserveCode(
                    session_id=self.olympiad_session_id,
                    class_parallel=parallel,
                    code=grade9_codes[code_index].code
                )
                self.session.add(reserve_code)
                code_index += 1
                total_allocated += 1

            self.stats[8]['reserve_codes'] += codes_to_allocate
            print(f"   ✅ {parallel}: выделено {codes_to_allocate} резервных кодов")

        print(f"   📊 Всего выделено резервных кодов: {total_allocated}")

    def _print_statistics(self):
        """Вывести итоговую статистику"""
        print("\n" + "=" * 80)
        print("📈 ИТОГОВАЯ СТАТИСТИКА")
        print("=" * 80)

        for class_number in sorted(self.stats.keys()):
            stats = self.stats[class_number]
            print(f"\n{class_number} класс:")
            print(f"  👥 Учеников: {stats['students']}")
            print(f"  ✅ Распределено кодов: {stats['codes_assigned']}")
            print(f"  📦 Осталось кодов: {stats['codes_available']}")
            if class_number == 8 and stats['reserve_codes'] > 0:
                print(f"  🔄 Резервных кодов (из 9 кл.): {stats['reserve_codes']}")


class ExcelExporter:
    """Экспорт распределенных кодов в Excel файлы"""

    def __init__(self, session: AsyncSession, olympiad_session_id: int, output_dir: str = "exports"):
        self.session = session
        self.olympiad_session_id = olympiad_session_id
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

    async def export_all(self):
        """Экспортировать файлы для всех классов"""
        print(f"\n📁 Экспортируем файлы в директорию: {self.output_dir}")
        print("=" * 80)

        # Получаем олимпиаду
        olympiad = await self.session.get(OlympiadSession, self.olympiad_session_id)

        # Получаем все классы с распределенными кодами
        result = await self.session.execute(
            select(OlympiadCode.class_number)
            .where(
                and_(
                    OlympiadCode.session_id == self.olympiad_session_id,
                    OlympiadCode.is_assigned == True
                )
            )
            .distinct()
        )
        classes = [row[0] for row in result.fetchall()]

        for class_number in sorted(classes):
            await self._export_class(class_number, olympiad)

        print(f"\n✅ Экспорт завершен! Файлы находятся в: {self.output_dir}")

    async def _export_class(self, class_number: int, olympiad: OlympiadSession):
        """Экспортировать файл для одного класса"""
        # Получаем учеников с кодами
        result = await self.session.execute(
            select(Student, OlympiadCode)
            .join(OlympiadCode, Student.id == OlympiadCode.student_id)
            .where(
                and_(
                    OlympiadCode.session_id == self.olympiad_session_id,
                    OlympiadCode.class_number == class_number,
                    OlympiadCode.is_assigned == True
                )
            )
            .order_by(Student.parallel, Student.full_name)
        )
        students_codes = result.fetchall()

        if not students_codes:
            return

        # Группируем по параллелям
        by_parallel = defaultdict(list)
        for student, code in students_codes:
            parallel = student.parallel or ''
            by_parallel[parallel].append((student, code))

        # Создаем файл для каждой параллели
        for parallel, data in sorted(by_parallel.items()):
            filename = f"{class_number}{parallel}.xlsx"
            filepath = os.path.join(self.output_dir, filename)

            await self._create_excel_file(filepath, class_number, parallel, data, olympiad)
            print(f"   ✅ Создан: {filename} ({len(data)} учеников)")

    async def _create_excel_file(self, filepath: str, class_number: int, parallel: str, data: list, olympiad):
        """Создать Excel файл для класса"""
        wb = Workbook()
        ws = wb.active
        ws.title = f"{class_number}{parallel}"

        # Стили
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Заголовок
        ws.merge_cells('A1:B1')
        ws['A1'] = f"Олимпиада: {olympiad.subject} | Класс: {class_number}{parallel}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = header_alignment

        ws.merge_cells('A2:B2')
        ws['A2'] = f"Дата: {olympiad.date.strftime('%d.%m.%Y')}"
        ws['A2'].alignment = header_alignment

        # Таблица с кодами
        ws['A4'] = "ФИО"
        ws['B4'] = "Код"

        for col in ['A4', 'B4']:
            ws[col].font = header_font
            ws[col].fill = header_fill
            ws[col].alignment = header_alignment
            ws[col].border = border

        # Данные
        row = 5
        for student, code in data:
            ws[f'A{row}'] = student.full_name
            ws[f'B{row}'] = code.code

            for col in [f'A{row}', f'B{row}']:
                ws[col].border = border

            row += 1

        # Для 8 класса добавляем резервные коды
        if class_number == 8:
            await self._add_reserve_codes_sheet(wb, parallel)

        # Авто-ширина колонок
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 30

        wb.save(filepath)

    async def _add_reserve_codes_sheet(self, workbook: Workbook, parallel: str):
        """Добавить лист с резервными кодами для 8 класса"""
        result = await self.session.execute(
            select(Grade8ReserveCode)
            .where(
                and_(
                    Grade8ReserveCode.session_id == self.olympiad_session_id,
                    Grade8ReserveCode.class_parallel == f"8{parallel}"
                )
            )
            .order_by(Grade8ReserveCode.id)
        )
        reserve_codes = result.scalars().all()

        if not reserve_codes:
            return

        # Создаем новый лист
        ws = workbook.create_sheet(title="Резервные коды")

        # Стили
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Заголовок
        ws.merge_cells('A1:B1')
        ws['A1'] = "Резервные коды (из пула 9 класса)"
        ws['A1'].font = Font(bold=True, size=14, color="FF6B6B")
        ws['A1'].alignment = header_alignment

        ws.merge_cells('A2:B2')
        ws['A2'] = "Эти коды выдаются по запросу любому ученику класса"
        ws['A2'].font = Font(italic=True, size=10)
        ws['A2'].alignment = header_alignment

        # Таблица
        ws['A4'] = "№"
        ws['B4'] = "Код"

        for col in ['A4', 'B4']:
            ws[col].font = header_font
            ws[col].fill = header_fill
            ws[col].alignment = header_alignment

        # Данные
        for i, code in enumerate(reserve_codes, 1):
            ws[f'A{i+4}'] = i
            ws[f'B{i+4}'] = code.code

        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 30


async def main():
    """Главная функция"""
    parser = argparse.ArgumentParser(description="Распределение кодов между учениками")
    parser.add_argument('--session-id', type=int, required=True, help='ID сессии олимпиады')
    parser.add_argument('--output-dir', type=str, default='exports', help='Директория для экспорта файлов')
    parser.add_argument('--skip-export', action='store_true', help='Пропустить экспорт файлов')
    args = parser.parse_args()

    # Получаем сессию БД
    async for session in get_async_session():
        try:
            # Распределяем коды
            distributor = CodeDistributor(session, args.session_id)
            await distributor.distribute_all()

            # Экспортируем файлы
            if not args.skip_export:
                exporter = ExcelExporter(session, args.session_id, args.output_dir)
                await exporter.export_all()

        except Exception as e:
            print(f"\n❌ Ошибка: {e}")
            import traceback
            traceback.print_exc()
            await session.rollback()
        finally:
            await session.close()


if __name__ == "__main__":
    asyncio.run(main())
