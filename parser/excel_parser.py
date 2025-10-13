import openpyxl
from typing import List, Dict
import logging
import re  # Добавьте этот импорт!

logger = logging.getLogger(__name__)


class StudentExcelParser:
    """Парсер Excel файлов со списком учеников"""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(file_path, data_only=True)
        
    def parse(self) -> List[Dict]:
        """
        Парсит Excel файл и возвращает список учеников
        
        Returns:
            List[Dict] с ключами: full_name, class_number
        """
        logger.info(f"Начинаем парсинг файла: {self.file_path}")
        
        students = []
        sheet = self.workbook.active
        
        # Ищем заголовки - УЛУЧШЕННАЯ ВЕРСИЯ
        header_row = None
        fio_col = None
        class_col = None
        
        # Расширенные варианты названий колонок
        fio_variants = ['фио', 'имя', 'ф.и.о', 'ф.и.о.', 'фамилия', 'студент', 'ученик']
        class_variants = ['класс', 'группа', 'grade', 'class']
        
        for row_idx, row in enumerate(sheet.iter_rows(max_row=15), start=1):  # Увеличили до 15 строк
            for col_idx, cell in enumerate(row, start=1):
                if not cell.value:
                    continue
                    
                cell_text = str(cell.value).lower().strip()
                
                # Проверяем варианты для ФИО
                if not fio_col:
                    for variant in fio_variants:
                        if variant in cell_text:
                            fio_col = col_idx
                            header_row = row_idx
                            logger.info(f"Найдена колонка ФИО: '{cell.value}' в столбце {col_idx}")
                            break
                
                # Проверяем варианты для Класса
                if not class_col:
                    for variant in class_variants:
                        if variant in cell_text:
                            class_col = col_idx
                            header_row = row_idx
                            logger.info(f"Найдена колонка Класс: '{cell.value}' в столбце {col_idx}")
                            break
            
            if fio_col and class_col:
                break
        
        if not fio_col or not class_col:
            # Выводим первые 5 строк для отладки
            logger.error("Не найдены колонки ФИО и Класс!")
            logger.error("Первые 5 строк файла:")
            for i, row in enumerate(sheet.iter_rows(max_row=5), 1):
                row_data = [str(cell.value) if cell.value else '' for cell in row[:10]]
                logger.error(f"Строка {i}: {row_data}")
            
            raise ValueError(
                "Не найдены колонки ФИО и Класс в таблице. "
                "Проверьте, что в первых 15 строках есть заголовки с названиями колонок. "
                "Допустимые варианты: ФИО/Имя/Фамилия для имен, Класс/Группа для класса."
            )
        
        logger.info(f"Найдены колонки: ФИО - столбец {fio_col}, Класс - столбец {class_col}")
        logger.info(f"Строка заголовка: {header_row}")
        
        # Остальной код без изменений...
        # Читаем данные
        for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row + 1), start=header_row + 1):
            fio_cell = row[fio_col - 1]
            class_cell = row[class_col - 1]
            
            if not fio_cell.value:
                continue
            
            full_name = str(fio_cell.value).strip()
            
            # Парсим номер класса
            class_value = str(class_cell.value).strip() if class_cell.value else ""
            
            try:
                # Извлекаем ПЕРВОЕ число из строки (для 8-Т2, 7-Т1 и т.д.)
                match = re.search(r'(\d+)', class_value)
                if not match:
                    logger.warning(f"Пропуск строки {row_idx}: не найдена цифра в '{class_value}'")
                    continue
                
                class_number = int(match.group(1))
                
                if not (4 <= class_number <= 11):
                    logger.warning(f"Пропуск строки {row_idx}: некорректный класс {class_value} (число {class_number})")
                    continue
                
                # Извлекаем параллель (все после первой цифры)
                parallel = class_value[match.end():].strip() if match.end() < len(class_value) else None
                if parallel and parallel.startswith('-'):
                    parallel = parallel[1:].strip()

                students.append({
                    "full_name": full_name,
                    "class_number": class_number,
                    "parallel": parallel if parallel else None,
                    "row_number": row_idx
                })
                
            except (ValueError, TypeError) as e:
                logger.warning(f"Пропуск строки {row_idx}: ошибка парсинга класса '{class_value}' - {e}")
                continue
        
        logger.info(f"Успешно распарсено {len(students)} учеников")
        
        return students
    
    def validate(self, students: List[Dict]) -> Dict:
        """
        Валидация распарсенных данных
        
        Returns:
            Dict с результатами валидации
        """
        issues = {
            "duplicates": [],
            "invalid_names": [],
            "class_distribution": {}
        }
        
        # Проверка дубликатов
        seen_names = {}
        for student in students:
            name = student["full_name"]
            if name in seen_names:
                issues["duplicates"].append({
                    "name": name,
                    "rows": [seen_names[name], student["row_number"]]
                })
            else:
                seen_names[name] = student["row_number"]
        
        # Проверка корректности имен
        for student in students:
            name = student["full_name"]
            # Минимальная проверка: имя должно содержать хотя бы 2 слова
            if len(name.split()) < 2:
                issues["invalid_names"].append({
                    "name": name,
                    "row": student["row_number"]
                })
        
        # Распределение по классам
        for student in students:
            class_num = student["class_number"]
            issues["class_distribution"][class_num] = issues["class_distribution"].get(class_num, 0) + 1
        
        return issues


def parse_students_excel(file_path: str) -> tuple[List[Dict], Dict]:
    """
    Удобная функция для парсинга Excel файла
    
    Args:
        file_path: путь к Excel файлу
        
    Returns:
        tuple: (список учеников, результаты валидации)
    """
    parser = StudentExcelParser(file_path)
    students = parser.parse()
    validation = parser.validate(students)
    
    return students, validation


if __name__ == "__main__":
    # Тестирование
    import sys
    
    if len(sys.argv) < 2:
        print("Использование: python excel_parser.py <файл.xlsx>")
        sys.exit(1)
    
    students, validation = parse_students_excel(sys.argv[1])
    
    print(f"\n✅ Найдено учеников: {len(students)}")
    print(f"\n📊 Распределение по классам:")
    for class_num in sorted(validation["class_distribution"].keys()):
        count = validation["class_distribution"][class_num]
        print(f"  {class_num} класс: {count} учеников")
    
    if validation["duplicates"]:
        print(f"\n⚠️ Найдено дубликатов: {len(validation['duplicates'])}")
        for dup in validation["duplicates"][:5]:
            print(f"  - {dup['name']} (строки {dup['rows']})")
    
    if validation["invalid_names"]:
        print(f"\n⚠️ Подозрительные имена: {len(validation['invalid_names'])}")
        for inv in validation["invalid_names"][:5]:
            print(f"  - {inv['name']} (строка {inv['row']})")