import openpyxl
from typing import List, Dict
import logging

logger = logging.getLogger(__name__)


class StudentExcelParser:
    """Парсер Excel файлов со списком учеников"""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(file_path, data_only=True)
        
    def parse(self) -> List[Dict]:
        """
        Парсит Excel файл и возвращает список учеников
        
        Returns:
            List[Dict] с ключами: full_name, class_number
        """
        logger.info(f"Начинаем парсинг файла: {self.file_path}")
        
        students = []
        sheet = self.workbook.active
        
        # Ищем заголовки
        header_row = None
        fio_col = None
        class_col = None
        
        for row_idx, row in enumerate(sheet.iter_rows(max_row=10), start=1):
            for col_idx, cell in enumerate(row, start=1):
                if cell.value:
                    cell_text = str(cell.value).lower().strip()
                    
                    if 'фио' in cell_text or 'имя' in cell_text:
                        fio_col = col_idx
                        header_row = row_idx
                    elif 'класс' in cell_text:
                        class_col = col_idx
                        header_row = row_idx
            
            if fio_col and class_col:
                break
        
        if not fio_col or not class_col:
            raise ValueError("Не найдены колонки ФИО и Класс в таблице")
        
        logger.info(f"Найдены колонки: ФИО - {fio_col}, Класс - {class_col}")
        logger.info(f"Строка заголовка: {header_row}")
        
        # Читаем данные
        for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row + 1), start=header_row + 1):
            fio_cell = row[fio_col - 1]
            class_cell = row[class_col - 1]
            
            if not fio_cell.value:
                continue
            
            full_name = str(fio_cell.value).strip()
            
            # Парсим номер класса
            class_value = str(class_cell.value).strip() if class_cell.value else ""
            
            try:
                # Извлекаем число из строки типа "8А", "9 Б", "11"
                class_number = int(''.join(filter(str.isdigit, class_value)))
                
                if not (4 <= class_number <= 11):
                    logger.warning(f"Пропуск строки {row_idx}: некорректный класс {class_value}")
                    continue
                
                students.append({
                    "full_name": full_name,
                    "class_number": class_number,
                    "row_number": row_idx
                })
                
            except (ValueError, TypeError) as e:
                logger.warning(f"Пропуск строки {row_idx}: ошибка парсинга класса '{class_value}' - {e}")
                continue
        
        logger.info(f"Успешно распарсено {len(students)} учеников")
        
        return students
    
    def validate(self, students: List[Dict]) -> Dict:
        """
        Валидация распарсенных данных
        
        Returns:
            Dict с результатами валидации
        """
        issues = {
            "duplicates": [],
            "invalid_names": [],
            "class_distribution": {}
        }
        
        # Проверка дубликатов
        seen_names = {}
        for student in students:
            name = student["full_name"]
            if name in seen_names:
                issues["duplicates"].append({
                    "name": name,
                    "rows": [seen_names[name], student["row_number"]]
                })
            else:
                seen_names[name] = student["row_number"]
        
        # Проверка корректности имен
        for student in students:
            name = student["full_name"]
            # Минимальная проверка: имя должно содержать хотя бы 2 слова
            if len(name.split()) < 2:
                issues["invalid_names"].append({
                    "name": name,
                    "row": student["row_number"]
                })
        
        # Распределение по классам
        for student in students:
            class_num = student["class_number"]
            issues["class_distribution"][class_num] = issues["class_distribution"].get(class_num, 0) + 1
        
        return issues


def parse_students_excel(file_path: str) -> tuple[List[Dict], Dict]:
    """
    Удобная функция для парсинга Excel файла
    
    Args:
        file_path: путь к Excel файлу
        
    Returns:
        tuple: (список учеников, результаты валидации)
    """
    parser = StudentExcelParser(file_path)
    students = parser.parse()
    validation = parser.validate(students)
    
    return students, validation


if __name__ == "__main__":
    # Тестирование
    import sys
    
    if len(sys.argv) < 2:
        print("Использование: python excel_parser.py <файл.xlsx>")
        sys.exit(1)
    
    students, validation = parse_students_excel(sys.argv[1])
    
    print(f"\n✅ Найдено учеников: {len(students)}")
    print(f"\n📊 Распределение по классам:")
    for class_num in sorted(validation["class_distribution"].keys()):
        count = validation["class_distribution"][class_num]
        print(f"  {class_num} класс: {count} учеников")
    
    if validation["duplicates"]:
        print(f"\n⚠️ Найдено дубликатов: {len(validation['duplicates'])}")
        for dup in validation["duplicates"][:5]:
            print(f"  - {dup['name']} (строки {dup['rows']})")
    
    if validation["invalid_names"]:
        print(f"\n⚠️ Подозрительные имена: {len(validation['invalid_names'])}")
        for inv in validation["invalid_names"][:5]:
            print(f"  - {inv['name']} (строка {inv['row']})")